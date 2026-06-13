#!/usr/bin/env python3
"""
ETL: Purple Star Luck workbook  ->  logical JSON data structures.

Imports ALL of Bill's authored content from the canonical workbook (plus the
Houses trait-bank) into three normalized read-models the report engine consumes:

  website/data/ps/stars.json       - the canonical 37-star registry (Bill's set,
                                      not iztro's 66), joined to hanzi/pinyin/
                                      modern-name via the existing overlay.
  website/data/ps/narratives.json  - all 8 narrative sheets, normalized.
  website/data/ps/fate.json        - per-palace x luck-level trait bank (Houses.docx),
                                      the v1 "fate" layer pending Bill's 5-field rewrite.

Run from repo root:  python3 docs/features/purple-star-report/etl_purple_star.py
"""
import json, os, re, subprocess, sys

ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", ".."))
XLSX = os.path.join(ROOT, "docs/architecture/readings/Purple-Star-Luck/Purple Star Luck - MergedUpdated.xlsx")
HOUSES = os.path.join(ROOT, "docs/architecture/readings/Purple-Star-Luck/Purple Star Houses.docx")
OUTDIR = os.path.join(ROOT, "website/data/ps")
# Existing iztro-based overlay -> source of hanzi / pinyin / modern names, keyed by legacy alias.
EXISTING_STARS = os.path.join(ROOT, "website/data/purple-star-stars.json")

import openpyxl

os.makedirs(OUTDIR, exist_ok=True)


def clean(v):
    if v is None:
        return None
    s = " ".join(str(v).split())
    return s if s and s.upper() != "NULL" else None


def rows(ws, wb):
    return list(wb[ws].iter_rows(values_only=True))


# Our 12 internal palace keys + the workbook's labels for them.
PALACE_KEYS = ["Ming", "Siblings", "Marriage", "Children", "Wealth", "Health",
               "Travel", "Associates", "Career", "Property", "Happiness", "Parents"]
SHEET_PALACE_TO_KEY = {
    "ming / fate": "Ming", "ming": "Ming",
    "siblings": "Siblings", "marriage": "Marriage", "children": "Children",
    "wealth": "Wealth", "health": "Health", "travel": "Travel",
    "associates": "Associates", "career": "Career", "property": "Property",
    "happiness": "Happiness", "parents": "Parents",
}
# Palace_Conclusion luck columns -> our category keys.
LUCK_COLS = {
    "Most Lucky": "mostLucky", "Least Lucky": "leastLucky",
    "Generally Lucky": "generallyLucky", "Generally Unlucky": "generallyUnlucky",
    "Mixed Luck": "mixed",
}


# ─────────────────────────────────────────────────────────────────────────────
# 1) STARS — the canonical 37 (Bill's system), joined to hanzi/pinyin/modern.
# ─────────────────────────────────────────────────────────────────────────────
def build_stars(wb):
    # Index the existing iztro-overlay by its Kwok Man-Ho legacy alias.
    existing = json.load(open(EXISTING_STARS))
    by_legacy = {}
    for s in existing:
        if s.get("legacyAlias"):
            by_legacy[s["legacyAlias"].strip().lower()] = s
    # Manual bridges where the workbook's English differs from the overlay alias.
    ALIASES = {
        "siren star": "bell star",
        "ringing/siren star": "bell star",
        "sheep blade star (destruction)": "sheep star (destruction)",
    }
    # The four Transformations are mutagens (a star property in iztro), not placed
    # stars; carry them as kind="Transformation" with the Si-Hua code.
    MUTAGEN = {
        "transforming authority": "Quan", "transforming salary": "Lu",
        "transforming examination": "Ke", "transforming jealousy": "Ji",
    }

    out = []
    unmatched = []
    for r in rows("Stars", wb)[1:]:
        sid, chinese, english, typ = (clean(r[0]), clean(r[1]), clean(r[2]), clean(r[3]))
        if not chinese:
            continue
        legacy = (english or "").strip().lower()
        rec = {
            "id": int(sid) if sid and sid.isdigit() else sid,
            "romanization": chinese,            # workbook's romanization, e.g. "Zi Wei"
            "legacyEnglish": english,           # Kwok Man-Ho name (hidden alias)
            "billType": typ,                    # Bill's Major/Minor split
            "hanzi": None, "pinyin": None, "iztroKey": None,
            "modernName": None, "nameStatus": "unnamed",
            "element": None, "courtRole": None,
            "mutagenCode": MUTAGEN.get(legacy),
            "kind": "Transformation" if legacy in MUTAGEN else typ,
        }
        # Stars iztro has no en-keyed catalog row for (e.g. 大耗) — manual hanzi.
        MANUAL = {"heavenly destroyer": {"hanzi": "大耗", "pinyin": "Dà Hào"}}
        match = by_legacy.get(legacy) or by_legacy.get(ALIASES.get(legacy, ""))
        if not match and legacy in MANUAL:
            rec["hanzi"] = MANUAL[legacy]["hanzi"]
            rec["pinyin"] = MANUAL[legacy]["pinyin"]
        if match:
            rec["hanzi"] = match.get("hanzi")
            rec["pinyin"] = match.get("pinyin")
            rec["iztroKey"] = match.get("iztroKey")
            rec["element"] = match.get("element")
            rec["courtRole"] = match.get("courtRole")
            if match.get("nameStatus") == "locked" and match.get("name"):
                rec["modernName"] = match["name"]
                rec["nameStatus"] = "locked"
            elif match.get("name"):
                rec["modernName"] = match["name"]
                rec["nameStatus"] = match.get("nameStatus", "draft")
        elif not rec["mutagenCode"]:
            unmatched.append(english)
        out.append(rec)
    if unmatched:
        print(f"  [stars] no hanzi join for: {unmatched}", file=sys.stderr)
    return out


# ─────────────────────────────────────────────────────────────────────────────
# 2) NARRATIVES — all 8 sheets, normalized into one object.
# ─────────────────────────────────────────────────────────────────────────────
def combo_key(parts):
    """Normalize a star-luck combo (up to 3 of VL/L/U/VUL) to a sorted key."""
    order = {"VL": 0, "L": 1, "U": 2, "VUL": 3}
    codes = [p for p in (clean(x) for x in parts) if p in order]
    codes.sort(key=lambda c: order[c])
    return ",".join(codes)


def build_narratives(wb):
    n = {}

    # Palace_Conclusion: palace x {5 luck categories}
    pc = {}
    ws = rows("Palace_Conclusion", wb)
    hdr = [clean(c) for c in ws[0]]
    for r in ws[1:]:
        label = clean(r[0])
        if not label:
            continue
        key = SHEET_PALACE_TO_KEY.get(label.lower())
        if not key:
            continue
        pc[key] = {}
        for ci, col in enumerate(hdr):
            if col in LUCK_COLS and clean(r[ci]):
                pc[key][LUCK_COLS[col]] = clean(r[ci])
    n["palaceConclusion"] = pc

    # Luckiest_Palcace: palace x {luckiest, unluckiest}
    lp = {}
    ws = rows("Luckiest_Palcace", wb)
    for r in ws[1:]:
        label = clean(r[0])
        key = SHEET_PALACE_TO_KEY.get((label or "").lower())
        if key:
            lp[key] = {"luckiest": clean(r[1]), "unluckiest": clean(r[2])}
    n["palaceExtreme"] = lp

    # Luckiest_Time_Period: decade# x {luckiest, unluckiest}
    tp = {}
    for r in rows("Luckiest_Time_Period", wb)[1:]:
        d = clean(r[0])
        if d and d.isdigit():
            tp[d] = {"luckiest": clean(r[1]), "unluckiest": clean(r[2])}
    n["decadePeriod"] = tp

    # Decade_Yun: decade# x yun-rating -> text
    dy = {}
    for r in rows("Decade_Yun", wb)[1:]:
        d, yun, txt = clean(r[0]), clean(r[1]), clean(r[2])
        if d and yun:
            dy.setdefault(d, {})[yun] = txt
    n["decadeYun"] = dy

    # Decade_Start_Narrative: auspiciousness -> {firstYear, yearsOfDecade}
    ds = {}
    ws = rows("Decade_Start_Narrative", wb)
    for r in ws[1:]:
        ausp = clean(r[0])
        if not ausp:
            continue
        ds[ausp] = {"firstYear": clean(r[1]) or clean(r[2]),
                    "yearsOfDecade": clean(r[3]) if len(r) > 3 else None}
    n["decadeStart"] = ds

    # Year_Descriptions: age-bucket label -> {5 luck categories}
    yd = []
    ws = rows("Year_Descriptions", wb)
    hdr = [clean(c) for c in ws[0]]
    catcols = {ci: LUCK_COLS[c] for ci, c in enumerate(hdr) if c in LUCK_COLS}
    for r in ws[1:]:
        bucket = clean(r[1]) if len(r) > 1 else None
        idx = clean(r[0])
        if not bucket:
            continue
        entry = {"index": idx, "bucket": bucket}
        for ci, cat in catcols.items():
            if clean(r[ci]):
                entry[cat] = clean(r[ci])
        yd.append(entry)
    n["yearDescriptions"] = yd

    # Major-Minor_Stars: the scoring engine. combo-key x scope x {major,minor}.
    ws = rows("Major-Minor_Stars", wb)
    hdr = [clean(c) for c in ws[0]]
    COL = {h: i for i, h in enumerate(hdr) if h}
    scope_map = {
        "decade": ("Major Star Narrative Decade", "Minor Star Narrative Decade"),
        "year": ("Major Star Narrative Year", "Minor Star Narrative Year"),
        "thisYear": ("Major Star - This Year", "Minor Star - This Year"),
        "month": ("Major Star Narrative Month", "Minor Star Narrative Month"),
        "palace": ("Palace Major Star", "Palace Minor Star"),
    }
    combos = {}
    for r in ws[1:]:
        key = combo_key(r[0:3])
        if not key:
            continue
        rec = combos.setdefault(key, {})
        for scope, (mc, nc) in scope_map.items():
            maj = clean(r[COL[mc]]) if mc in COL and COL[mc] < len(r) else None
            mino = clean(r[COL[nc]]) if nc in COL and COL[nc] < len(r) else None
            if maj or mino:
                rec[scope] = {"major": maj, "minor": mino}
    n["starCombo"] = combos

    # Next Year: luck rating -> {firstMonths, secondMonths}
    ny = {}
    for r in rows("Next Year", wb)[1:]:
        rating = clean(r[0])
        if rating:
            ny[rating] = {"firstMonths": clean(r[1]) if len(r) > 1 else None,
                          "secondMonths": clean(r[2]) if len(r) > 2 else None}
    n["nextYear"] = ny

    return n


# ─────────────────────────────────────────────────────────────────────────────
# 3) FATE — Houses.docx per-palace x luck-level trait bank (v1 fate layer).
# ─────────────────────────────────────────────────────────────────────────────
def build_fate():
    txt = subprocess.run(["textutil", "-convert", "txt", "-stdout", HOUSES],
                         capture_output=True, text=True).stdout
    lines = [l.replace("\t", " ").rstrip() for l in txt.splitlines()]
    # Houses uses descriptive section labels, not our palace keys.
    HEADER_ALIASES = {
        "ming": "Ming", "brothers and sisters": "Siblings", "marital": "Marriage",
        "man and woman": "Children", "wealth": "Wealth", "sickness": "Health",
        "moving": "Travel", "servants": "Associates", "officials": "Career",
        "property": "Property", "happiness": "Happiness", "fortune": "Happiness",
        "parents": "Parents",
    }
    level_re = re.compile(r"^(Very Lucky|Lucky|Mixed Luck|Mixed|Unlucky|Very Unlucky|VL|VUL|UL|L)\s*:\s*(.*)$", re.I)
    LEVEL = {"very lucky": "VL", "vl": "VL", "lucky": "L", "l": "L",
             "mixed": "mixed", "mixed luck": "mixed",
             "unlucky": "UL", "ul": "UL", "very unlucky": "VUL", "vul": "VUL"}

    def header_key(s):
        low = s.lower()
        if low.startswith("note"):
            return None
        for alias, key in HEADER_ALIASES.items():
            if low.startswith(alias) and len(s) < 70 and ":" not in s:
                return key
        return None

    fate = {}
    cur = None
    for line in lines:
        s = line.strip()
        if not s:
            continue
        if line.startswith("NOTE") or s.startswith("NOTE"):
            continue
        hk = header_key(s)
        if hk:
            cur = hk
            fate.setdefault(cur, {})
            continue
        lm = level_re.match(s)
        if lm and cur:
            lvl = LEVEL.get(lm.group(1).strip().lower())
            body = lm.group(2).strip()
            if lvl and body:
                fate[cur].setdefault(lvl, [])
                # split the trait blob into phrases
                phrases = [p.strip() for p in re.split(r"[.;]\s+", body) if p.strip()]
                fate[cur][lvl].extend(phrases)
    return fate


def main():
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    stars = build_stars(wb)
    narr = build_narratives(wb)
    fate = build_fate()

    json.dump(stars, open(os.path.join(OUTDIR, "stars.json"), "w"), ensure_ascii=False, indent=2)
    json.dump(narr, open(os.path.join(OUTDIR, "narratives.json"), "w"), ensure_ascii=False, indent=2)
    json.dump(fate, open(os.path.join(OUTDIR, "fate.json"), "w"), ensure_ascii=False, indent=2)

    # Report
    print(f"stars.json       : {len(stars)} stars "
          f"({sum(1 for s in stars if s['billType']=='Major')} major / "
          f"{sum(1 for s in stars if s['billType']=='Minor')} minor), "
          f"{sum(1 for s in stars if s['hanzi'])} joined to hanzi")
    print("narratives.json  : " + ", ".join(f"{k}={len(v)}" for k, v in narr.items()))
    print(f"fate.json        : {len(fate)} palaces, "
          f"{sum(len(v) for v in fate.values())} palace×level trait banks")


if __name__ == "__main__":
    main()
