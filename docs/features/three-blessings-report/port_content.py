#!/usr/bin/env python3
"""
Three Blessings content porter (Phase 1).

Reads Bill's authored workbook
  docs/architecture/readings/ThreeBlessings/Three Blessing.xlsx        (primary, 19 tables)
  docs/architecture/readings/ThreeBlessings/Three Blessing - Final.xls (superset: + Improve tables)
and emits one faithful JSON file per lookup table into website/data/tb/.

This is the Three Blessings analogue of docs/features/four-pillars-report/port_content.py.
Phase 1 is a LOSSLESS extraction: each table becomes an array of row objects keyed by the
sheet's own header names (empty spacer columns dropped, numeric ids/LuckValueIDs coerced to
int). The engine phase will build keyed indices from these arrays. Intended lookup keys per
table are documented in PHASE-1-FINDINGS.md.

Run from repo root:  python3 docs/features/three-blessings-report/port_content.py
Requires: openpyxl (xlsx), a one-time soffice conversion of the .xls for the Improve tables.
"""
import json, os, re, sys, warnings, subprocess, tempfile
warnings.simplefilter("ignore")
import openpyxl

ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", ".."))
SRC_DIR = os.path.join(ROOT, "docs", "architecture", "readings", "ThreeBlessings")
PRIMARY = os.path.join(SRC_DIR, "Three Blessing.xlsx")
FINAL_XLS = os.path.join(SRC_DIR, "Three Blessing - Final.xls")
OUT_DIR = os.path.join(ROOT, "website", "data", "tb")
GOLD_DIR = os.path.join(ROOT, "docs", "features", "three-blessings-report", "golden")

# Sheets that live only in the "Final" superset workbook (the "Improve Your Luck" tables).
FINAL_ONLY = ["TBImproveElement", "TBImproveYinYang", "TBHarmony"]  # TBConflict is empty -> skipped


def kebab(sheet):
    s = re.sub(r"^TB", "", sheet)
    s = re.sub(r"(?<!^)(?=[A-Z])", "-", s).lower()
    return "tb-" + s


def coerce(v):
    if v is None:
        return None
    if isinstance(v, float) and v.is_integer():
        return int(v)
    if isinstance(v, str):
        return v.strip()
    return v


def extract_sheet(ws):
    """Return (rows, kept_headers). Drops columns whose header AND every cell are empty."""
    grid = [[coerce(c) for c in r] for r in ws.iter_rows(values_only=True)]
    if not grid:
        return [], []
    header = grid[0]
    body = grid[1:]
    ncol = len(header)
    keep = []
    for j in range(ncol):
        name = header[j]
        has_name = name not in (None, "")
        has_data = any(j < len(r) and r[j] not in (None, "") for r in body)
        if has_name and has_data:
            keep.append((j, str(name)))
        elif has_name and not has_data:
            keep.append((j, str(name)))  # keep named-but-empty (documents the schema)
        # unnamed spacer columns (no header) are dropped
    rows = []
    for r in body:
        if all(j >= len(r) or r[j] in (None, "") for j, _ in keep):
            continue  # fully blank row
        rows.append({name: (r[j] if j < len(r) else None) for j, name in keep})
    return rows, [n for _, n in keep]


def main():
    os.makedirs(OUT_DIR, exist_ok=True)
    os.makedirs(GOLD_DIR, exist_ok=True)

    wb = openpyxl.load_workbook(PRIMARY, read_only=True, data_only=True)
    primary_data = {}
    manifest = []
    for ws in wb.worksheets:
        rows, headers = extract_sheet(ws)
        primary_data[ws.title] = rows
        fname = kebab(ws.title) + ".json"
        with open(os.path.join(OUT_DIR, fname), "w") as f:
            json.dump(rows, f, ensure_ascii=False, indent=1)
        manifest.append((ws.title, fname, len(rows), headers, "Three Blessing.xlsx"))

    # Improve tables from the Final superset (convert .xls -> .xlsx once via soffice).
    final_data = {}
    with tempfile.TemporaryDirectory() as td:
        subprocess.run(["soffice", "--headless", "--convert-to", "xlsx", "--outdir", td, FINAL_XLS],
                       check=True, capture_output=True)
        conv = os.path.join(td, "Three Blessing - Final.xlsx")
        fwb = openpyxl.load_workbook(conv, data_only=True)
        for title in FINAL_ONLY:
            if title not in fwb.sheetnames:
                print(f"  WARN: {title} not in Final workbook", file=sys.stderr)
                continue
            rows, headers = extract_sheet(fwb[title])
            fname = kebab(title) + ".json"
            with open(os.path.join(OUT_DIR, fname), "w") as f:
                json.dump(rows, f, ensure_ascii=False, indent=1)
            manifest.append((title, fname, len(rows), headers, "Three Blessing - Final.xls"))
        # Diff shared sheets: primary vs Final (fidelity/provenance check).
        diffs = []
        for title in fwb.sheetnames:
            if title in primary_data:
                frows, _ = extract_sheet(fwb[title])
                if frows != primary_data[title]:
                    # count differing rows for the report
                    n = sum(1 for a, b in zip(frows, primary_data[title]) if a != b)
                    n += abs(len(frows) - len(primary_data[title]))
                    diffs.append((title, len(primary_data[title]), len(frows), n))

    # Golden samples: the two authored prototype reports -> plain text fixtures.
    import zipfile
    def docx_text(path):
        z = zipfile.ZipFile(path)
        xml = z.read("word/document.xml").decode("utf-8", "ignore")
        xml = re.sub(r"</w:p>", "\n", xml)
        txt = re.sub(r"<[^>]+>", "", xml)
        txt = txt.replace("&lt;", "<").replace("&gt;", ">").replace("&amp;", "&")
        return re.sub(r"\n{3,}", "\n\n", txt).strip()

    goldens = {
        "golden-david-1972.txt": "Three Blessings Prototype.docx",
        "golden-bill-1947.txt": "Three Blessings Prototype - Jan 2013.docx",
    }
    for out, src in goldens.items():
        with open(os.path.join(GOLD_DIR, out), "w") as f:
            f.write(docx_text(os.path.join(SRC_DIR, src)))

    # Report
    print("=== EMITTED TABLES (website/data/tb/) ===")
    total = 0
    for title, fname, n, headers, srcwb in sorted(manifest):
        total += n
        print(f"  {fname:32s} {n:4d} rows  <- {title}")
    print(f"  {'TOTAL':32s} {total:4d} rows across {len(manifest)} tables")
    print("\n=== PRIMARY vs FINAL diff (shared sheets) ===")
    if not diffs:
        print("  (no differences — primary and Final are identical on all shared sheets)")
    else:
        for title, pn, fn, n in diffs:
            print(f"  {title}: primary={pn} rows, final={fn} rows, differing_rows={n}")
    print("\n=== GOLDEN FIXTURES (docs/features/three-blessings-report/golden/) ===")
    for out in goldens:
        p = os.path.join(GOLD_DIR, out)
        print(f"  {out}  ({os.path.getsize(p)} bytes)")


if __name__ == "__main__":
    main()
