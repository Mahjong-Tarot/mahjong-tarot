#!/usr/bin/env python3
"""
Build the auspiciousness matrix template for Bill:
  rows  = the stars (our core 37 first, then the rest of the ~115 universe)
  cols  = the 12 palaces
  cells = 1..4 (least → most auspicious), chosen from a dropdown.

Output: working_files/purple-star-auspiciousness-matrix.xlsx (+ .csv)
Import the .xlsx (or .csv) straight into Google Sheets.
"""
import csv, json, os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", ".."))
STARS = json.load(open(os.path.join(ROOT, "website/data/ps/stars.json")))
OUT = os.path.join(ROOT, "working_files")

PALACES = ["Fate (Ming)", "Siblings", "Marriage", "Children", "Wealth", "Health",
           "Travel", "Friends", "Career", "Property", "Wellbeing", "Parents"]
HEAD = ["ID", "Star (English)", "Hanzi", "Pinyin", "Type", "Gloss / note"] + PALACES

# ── core 37 (our system) — Bill's English names + hanzi + pinyin ──────────────
core = []
for s in STARS:
    core.append([s["id"], s["legacyEnglish"] or s["romanization"], s["hanzi"] or "",
                 s["pinyin"] or "", s["billType"], ""])

core_hanzi = {s["hanzi"] for s in STARS if s["hanzi"]}

# ── the rest of the ~115 universe (hanzi + short gloss; Bill writes his English) ──
# Static reference list (NOT generated from iztro). Glosses are placeholders.
EXTRA = [
    # adjective / 杂曜
    ("咸池", "Bath / Peach Blossom (romance, indulgence)"), ("天空", "Sky Void (emptiness)"),
    ("天虚", "Heavenly Hollow (depletion)"), ("天哭", "Heavenly Weeping (grief)"),
    ("华盖", "Canopy (solitude, spirituality)"), ("孤辰", "Lonesome (isolation)"),
    ("寡宿", "Widowed (isolation)"), ("阴煞", "Hidden Harm"), ("蜚廉", "Slander / gossip"),
    ("破碎", "Fragmentation"), ("天月", "Illness"), ("天伤", "Wounding"), ("天使", "Envoy / omen"),
    ("解神", "Dissolver (relief)"), ("天厨", "Heavenly Kitchen (food, comfort)"),
    ("天巫", "Shaman (intuition, promotion)"), ("天才", "Talent"), ("天寿", "Longevity"),
    ("天福", "Blessing"), ("天德", "Heaven Virtue"), ("月德", "Moon Virtue"),
    ("天官", "Office / rank"), ("天贵", "Nobility"), ("恩光", "Grace / favour"),
    ("台辅", "Honour"), ("封诰", "Patent / title"), ("八座", "Eight Seats (rank)"),
    ("三台", "Three Terraces (rank)"), ("凤阁", "Phoenix Pavilion (refinement)"),
    ("龙池", "Dragon Pool (talent)"), ("年解", "Annual Dissolver"),
    ("截路", "Cut-off (obstruction)"), ("旬空", "Cyclic Void"), ("空亡", "Void"),
    # 长生十二神 (changsheng)
    ("长生", "Birth (growth)"), ("沐浴", "Bathing (instability)"), ("冠带", "Capping (maturing)"),
    ("临官", "Ascending Office"), ("帝旺", "Imperium / Peak vitality"), ("衰", "Decline"),
    ("病", "Illness"), ("死", "Death"), ("墓", "Tomb / storage"), ("绝", "Severance"),
    ("胎", "Conception"), ("养", "Nurture"),
    # 博士十二神 (boshi)
    ("博士", "Scholar"), ("力士", "Strongman"), ("青龙", "Azure Dragon (luck)"),
    ("小耗", "Lesser Loss"), ("将军", "General"), ("奏书", "Memorial (documents)"),
    ("飞廉", "Flying Slander"), ("喜神", "Joy Spirit"), ("病符", "Sickness Token"),
    ("伏兵", "Ambush"), ("官府", "Officialdom / litigation"),
    # 岁前十二神 (suiqian)
    ("岁建", "Year Establish"), ("晦气", "Gloom"), ("丧门", "Mourning Gate"),
    ("贯索", "Entangling Rope"), ("官符", "Official Reprimand"), ("龙德", "Dragon Virtue"),
    ("白虎", "White Tiger (conflict)"), ("吊客", "Condolence Visitor"),
    # 将前十二神 (jiangqian)
    ("将星", "Commander Star"), ("攀鞍", "Mounting Saddle"), ("岁驿", "Year Post (travel)"),
    ("息神", "Resting Spirit"), ("劫煞", "Robbery"), ("灾煞", "Calamity"),
    ("天煞", "Heaven Harm"), ("指背", "Backbiting"), ("亡神", "Loss Spirit"), ("月煞", "Month Harm"),
]
extra = []
nid = 38
seen = set(core_hanzi)
for hz, gloss in EXTRA:
    if hz in seen:
        continue
    seen.add(hz)
    extra.append([nid, "", hz, "", "Additional", gloss])
    nid += 1

rows = core + extra

# ── write CSV ─────────────────────────────────────────────────────────────────
os.makedirs(OUT, exist_ok=True)
with open(os.path.join(OUT, "purple-star-auspiciousness-matrix.csv"), "w", newline="") as f:
    w = csv.writer(f)
    w.writerow(HEAD)
    for r in rows:
        w.writerow(r + [""] * len(PALACES))

# ── write XLSX (frozen header, dropdowns 1–4, core highlighted) ────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Auspiciousness"

instr = ("Rate each star 1–4 in each palace: 1 = least auspicious, 4 = most auspicious. "
         "Leave blank if a star is not used. Core 37 (your system) are shaded; the rest fill toward ~115.")
ws.append([instr])
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEAD))
ws["A1"].font = Font(italic=True, color="555555")
ws["A1"].alignment = Alignment(wrap_text=False)

ws.append(HEAD)
hdr_fill = PatternFill("solid", fgColor="2A1C47")
for c in ws[2]:
    c.font = Font(bold=True, color="F5EDDA")
    c.fill = hdr_fill
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

core_fill = PatternFill("solid", fgColor="F6ECCB")
for r in rows:
    ws.append(r + [""] * len(PALACES))
    if r[4] != "Additional":
        for c in ws[ws.max_row]:
            c.fill = core_fill

# dropdown 1–4 on the 12 palace columns
dv = DataValidation(type="list", formula1='"1,2,3,4"', allow_blank=True)
ws.add_data_validation(dv)
first_pcol = 7  # column G
data_start = 3
data_end = ws.max_row
from openpyxl.utils import get_column_letter
for ci in range(first_pcol, first_pcol + len(PALACES)):
    col = get_column_letter(ci)
    dv.add(f"{col}{data_start}:{col}{data_end}")
    ws.column_dimensions[col].width = 11

ws.column_dimensions["A"].width = 5
ws.column_dimensions["B"].width = 28
ws.column_dimensions["C"].width = 7
ws.column_dimensions["D"].width = 12
ws.column_dimensions["E"].width = 11
ws.column_dimensions["F"].width = 34
ws.freeze_panes = "G3"

wb.save(os.path.join(OUT, "purple-star-auspiciousness-matrix.xlsx"))
print(f"rows: {len(rows)} ({len(core)} core + {len(extra)} additional) × {len(PALACES)} palaces")
print("wrote working_files/purple-star-auspiciousness-matrix.{xlsx,csv}")
