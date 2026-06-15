#!/usr/bin/env python3
"""
Generate website/data/purple-star-palaces.json — Bill's authored 12-palace
narratives for the Purple Star report (PR3).

Source: docs/architecture/readings/Purple-Star-Luck/Purple Star Luck - MergedUpdated.xlsx
  - sheet Palace_Conclusion : per palace × {Most/Least/Generally Lucky, Generally Unlucky, Mixed}
  - sheet Luckiest_Palcace  : per palace × {Luckiest, Unluckiest}   (when this is your single best/worst palace)

Keyed by our internal palace name (the PALACE_LABELS values in lib/purpleStar.js):
Ming, Siblings, Marriage, Children, Wealth, Health, Travel, Servants, Career,
Property, Leisure, Parents.

All text is Bill's — this script only reshapes it. The DB table
purple_star_narratives is the canonical store; this JSON is the read-model.
"""
import json
import openpyxl

XLSX = "docs/architecture/readings/Purple-Star-Luck/Purple Star Luck - MergedUpdated.xlsx"

# Sheet palace label -> our internal palace key
SHEET_TO_KEY = {
    "Ming / Fate": "Ming", "Ming": "Ming",
    "Siblings": "Siblings",
    "Marriage": "Marriage",
    "Children": "Children",
    "Wealth": "Wealth",
    "Health": "Health",
    "Travel": "Travel",
    "Associates": "Servants",
    "Career": "Career",
    "Property": "Property",
    "Happiness": "Leisure",
    "Parents": "Parents",
}


def clean(v):
    if v is None:
        return None
    return " ".join(str(v).split())  # collapse whitespace/newlines


wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
out = {}

# Palace_Conclusion: Palace | Most Lucky | Least Lucky | Generally Lucky | Generally Unlucky | Mixed Luck
for r in wb["Palace_Conclusion"].iter_rows(min_row=2, values_only=True):
    label = clean(r[0])
    if not label or label not in SHEET_TO_KEY:
        continue
    key = SHEET_TO_KEY[label]
    out.setdefault(key, {})["conclusion"] = {
        "mostLucky":       clean(r[1]),
        "leastLucky":      clean(r[2]),
        "generallyLucky":  clean(r[3]),
        "generallyUnlucky": clean(r[4]),
        "mixed":           clean(r[5]),
    }

# Luckiest_Palcace: Palace | Luckiest Description | Unluckiest Description
for r in wb["Luckiest_Palcace"].iter_rows(min_row=2, values_only=True):
    label = clean(r[0])
    if not label or label not in SHEET_TO_KEY:
        continue
    key = SHEET_TO_KEY[label]
    out.setdefault(key, {})["luckiest"] = clean(r[1])
    out[key]["unluckiest"] = clean(r[2])

# Validate: every palace has all fields
EXPECT = ["Ming", "Siblings", "Marriage", "Children", "Wealth", "Health",
          "Travel", "Servants", "Career", "Property", "Leisure", "Parents"]
problems = []
for k in EXPECT:
    if k not in out:
        problems.append(f"missing palace {k}")
        continue
    c = out[k].get("conclusion", {})
    for f in ["mostLucky", "leastLucky", "generallyLucky", "generallyUnlucky", "mixed"]:
        if not c.get(f):
            problems.append(f"{k}.conclusion.{f} empty")
    if not out[k].get("luckiest") or not out[k].get("unluckiest"):
        problems.append(f"{k} missing luckiest/unluckiest")

path = "website/data/purple-star-palaces.json"
with open(path, "w", encoding="utf-8") as f:
    json.dump(out, f, ensure_ascii=False, indent=2)
    f.write("\n")


# --- Seed SQL for the canonical DB store (purple_star_narratives) ---
def sql_str(v):
    return "null" if v is None else "'" + str(v).replace("'", "''") + "'"


CONCL_KEYS = ["mostLucky", "leastLucky", "generallyLucky", "generallyUnlucky", "mixed"]
rows = []
for key in EXPECT:
    for ck in CONCL_KEYS:
        rows.append(("palace_conclusion", key, ck, "", out[key]["conclusion"][ck]))
    rows.append(("palace_luckiest", key, "luckiest", "", out[key]["luckiest"]))
    rows.append(("palace_luckiest", key, "unluckiest", "", out[key]["unluckiest"]))

values = ",\n".join(
    "  (%s, %s, %s, %s, %s)" % (sql_str(c), sql_str(k1), sql_str(k2), sql_str(s), sql_str(t))
    for (c, k1, k2, s, t) in rows
)
sql = (
    "-- ============================================================\n"
    "-- Mahjong Tarot: seed Purple Star palace narratives\n"
    "-- Migration: 047_seed_purple_star_narratives\n"
    "-- Run this in Supabase SQL Editor (https://supabase.com/dashboard)\n"
    "--\n"
    "-- %d rows (12 palaces x 5 conclusion categories + luckiest/unluckiest).\n"
    "-- Generated from purple-star-palaces.json by generate_palaces.py.\n"
    "-- Idempotent: re-running refreshes text via ON CONFLICT DO UPDATE.\n"
    "-- ============================================================\n\n"
    "insert into public.purple_star_narratives (category, key1, key2, scope, text) values\n"
    "%s\n"
    "on conflict (category, key1, key2, scope) do update set text = excluded.text;\n"
) % (len(rows), values)

sql_path = "website/supabase/047_seed_purple_star_narratives.sql"
with open(sql_path, "w", encoding="utf-8") as f:
    f.write(sql)

print(f"wrote {len(out)} palaces -> {path}")
print(f"wrote {len(rows)} narrative rows -> {sql_path}")
print("PROBLEMS:", problems if problems else "NONE")
